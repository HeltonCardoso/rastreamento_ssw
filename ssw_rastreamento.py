"""
Sistema de Rastreamento SSW - Versão Final com Detecção de Devoluções
Suporte completo a JSON/XML e classificação inteligente de status
"""

from gspread import worksheet
import requests
import pandas as pd
from datetime import datetime
import time
import os
import re
import threading
import json
from typing import Dict, List, Optional, Any
from concurrent.futures import ThreadPoolExecutor, as_completed
import logging
from pathlib import Path
import warnings
warnings.filterwarnings('ignore')

# ============================================
# CONFIGURAÇÕES
# ============================================

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler('ssw_rastreamento.log', encoding='utf-8'),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)

# Cores para Excel
CORES = {
    'ENTREGUE':    'E3F2E3',      # Verde claro
    'DEVOLVIDO':   'FFB6B6',      # Vermelho claro
    'ATRASADO':    'FFE0E0',      # Vermelho mais claro
    'ALERTA_1DIA': 'FFD8B1',      # Laranja claro
    'ALERTA_2DIAS': 'FFE5B4',     # Laranja mais claro
    'ALERTA_3DIAS': 'FFF3CD',     # Amarelo
    'SEM_DADOS':   'F0F0F0',      # Cinza
    'ERRO':        'FFCCCC',      # Vermelho erro
    'PADRAO':      'FFFFFF'       # Branco
}

# ============================================
# PROCESSADOR SSW
# ============================================

class ProcessadorSSW:
    API_URL = "https://ssw.inf.br/api/trackingdanfe"

    def __init__(
        self,
        delay_consultas: float = 0.0,
        workers: int = 5,
        timeout: int = 30,
        max_retries: int = 2,
        callback_log=None
    ):
        self.delay = delay_consultas
        self.workers = workers
        self.timeout = timeout
        self.max_retries = max_retries
        self.callback_log = callback_log

        self.headers = {
            'Content-Type': 'application/json',
            'Accept': 'application/json',
            'User-Agent': 'SSWRastreamento/3.0'
        }

        self._lock = threading.Lock()
        self.total_consultas = 0
        self.total_erros = 0
        self.total_sem_dados = 0
        self.total_sucessos = 0
        self._local = threading.local()

    def _session(self) -> requests.Session:
        if not hasattr(self._local, 'session'):
            s = requests.Session()
            s.headers.update(self.headers)
            adapter = requests.adapters.HTTPAdapter(
                pool_connections=20,
                pool_maxsize=20,
                max_retries=2
            )
            s.mount('https://', adapter)
            self._local.session = s
        return self._local.session

    def _log(self, mensagem: str, nivel: str = "info"):
        if self.callback_log:
            self.callback_log(mensagem, nivel)
        else:
            if nivel == "erro":
                logger.error(mensagem)
            elif nivel == "aviso":
                logger.warning(mensagem)
            else:
                logger.info(mensagem)

    def _inc(self, campo: str):
        with self._lock:
            setattr(self, campo, getattr(self, campo) + 1)

    def extrair_chave_nfe(self, texto) -> Optional[str]:
        """Extrai chave de 44 dígitos de diferentes formatos"""
        if pd.isna(texto) or not texto:
            return None
        
        texto = str(texto).strip()
        
        # Caso 1: URL completa do tracking
        if 'ssw.inf.br/app/tracking/' in texto:
            chave = texto.split('tracking/')[-1].split('?')[0].split('#')[0]
            if len(chave) >= 44:
                return chave[:44]
        
        # Caso 2: Apenas números
        numeros = re.sub(r'[^0-9]', '', texto)
        if len(numeros) == 44:
            return numeros
        
        # Caso 3: Regex para 44 dígitos consecutivos
        match = re.search(r'\d{44}', texto)
        if match:
            return match.group(0)
        
        return None

    def ler_planilha(self, caminho_arquivo: str, coluna_xml: str) -> pd.DataFrame:
        """Lê planilha e extrai chaves"""
        self._log(f"📂 Lendo arquivo: {caminho_arquivo}")
        ext = Path(caminho_arquivo).suffix.lower()
        
        try:
            if ext == '.csv':
                df = pd.read_csv(caminho_arquivo, encoding='utf-8-sig')
            elif ext in ['.xlsx', '.xls']:
                df = pd.read_excel(caminho_arquivo)
            else:
                raise ValueError(f"Formato não suportado: {ext}")

            if coluna_xml not in df.columns:
                raise ValueError(f"Coluna '{coluna_xml}' não encontrada. Disponíveis: {df.columns.tolist()}")

            df['chave_nfe'] = df[coluna_xml].apply(self.extrair_chave_nfe)
            total = len(df)
            validas = df['chave_nfe'].notna().sum()
            self._log(f"✅ Chaves válidas: {validas}/{total}")
            
            # Mostra amostra das primeiras chaves
            for i, row in df.head(3).iterrows():
                if pd.notna(row['chave_nfe']):
                    self._log(f"   Exemplo {i+1}: {row['chave_nfe']}", "debug")
            
            return df.dropna(subset=['chave_nfe'])
            
        except Exception as e:
            self._log(f"❌ Erro ao ler planilha: {e}", "erro")
            raise

    def extrair_previsao_entrega(self, eventos: List[Dict]) -> tuple:
        """
        Extrai a data de previsão dos eventos
        Retorna: (data_previsao_str, dias_restantes)
        """
        previsao_str = ""
        
        # Procura nos eventos por "Previsao de entrega"
        for evento in eventos:
            descricao = evento.get('descricao', '')
            if 'Previsao de entrega' in descricao:
                match = re.search(r'Previsao de entrega: (\d{2}/\d{2}/\d{2})', descricao)
                if match:
                    data_str = match.group(1)  # "20/02/26"
                    # Converte de DD/MM/AA para DD/MM/YYYY
                    partes = data_str.split('/')
                    if len(partes) == 3 and len(partes[2]) == 2:
                        ano = 2000 + int(partes[2])
                        previsao_str = f"{partes[0]}/{partes[1]}/{ano}"
                    else:
                        previsao_str = data_str
                    break
        
        # Calcula dias restantes
        dias_restantes = None
        if previsao_str:
            try:
                data_prev = datetime.strptime(previsao_str, '%d/%m/%Y')
                hoje = datetime.now()
                dias_restantes = (data_prev - hoje).days
            except Exception as e:
                self._log(f"Erro ao calcular dias: {e}", "debug")
        
        return previsao_str, dias_restantes

    def classificar_status(self, eventos: List[Dict]) -> Dict:
        """
        Classifica o status baseado nos eventos de rastreamento
        Retorna: {status, resumo, recomendacao, cor, prioridade}
        """
        if not eventos:
            return {
                'status': 'AGUARDANDO RASTREIO',
                'status_resumo': 'AGUARDANDO',
                'recomendacao': 'Sem rastreio na SSW. Verificar manualmente.',
                'cor': CORES['SEM_DADOS'],
                'prioridade': 4
            }
        
        # Verifica cada evento para identificar situação final
        entregue = False
        devolvido = False
        data_entrega = ""
        ultimo_status = ""
        
        for evento in eventos:
            ocorrencia = evento.get('ocorrencia', '').upper()
            descricao = evento.get('descricao', '').upper()
            
            # DETECÇÃO DE DEVOLUÇÃO (prioridade máxima)
            if any(palavra in ocorrencia for palavra in ['DEVOLVIDA', 'DEVOLUÇÃO', 'MERCADORIA DEVOLVIDA']):
                devolvido = True
                self._log(f"   🔄 Devolução detectada: {ocorrencia}", "debug")
                break  # Devolução é status final, não precisa verificar mais
            
            # DETECÇÃO DE ENTREGA
            if 'ENTREGUE' in ocorrencia or 'ENTREGA REALIZADA' in descricao:
                entregue = True
                data_entrega = evento.get('data_hora', '').split('T')[0]  # Pega apenas a data
                self._log(f"   ✅ Entrega detectada: {ocorrencia}", "debug")
            
            ultimo_status = ocorrencia
        
        # CASO 1: DEVOLVIDO (prioridade máxima)
        if devolvido:
            return {
                'status': 'DEVOLVIDO AO REMETENTE',
                'status_resumo': 'DEVOLVIDO',
                'recomendacao': '❌ Produto devolvido. Contatar cliente e transportadora para reenvio.',
                'cor': CORES['DEVOLVIDO'],
                'prioridade': 1,
                'data_entrega': data_entrega
            }
        
        # CASO 2: ENTREGUE
        if entregue:
            return {
                'status': 'ENTREGUE',
                'status_resumo': 'ENTREGUE',
                'recomendacao': '✅ Entrega realizada. Atualizar Intelipost e dar baixa no sistema.',
                'cor': CORES['ENTREGUE'],
                'prioridade': 5,
                'data_entrega': data_entrega
            }
        
        # CASO 3: EM TRÂNSITO (com ou sem previsão)
        previsao_str, dias_restantes = self.extrair_previsao_entrega(eventos)
        
        if dias_restantes is None:
            return {
                'status': 'EM TRÂNSITO',
                'status_resumo': 'TRÂNSITO',
                'recomendacao': 'Em trânsito, aguardando atualização de previsão',
                'cor': CORES['PADRAO'],
                'prioridade': 4,
                'previsao': previsao_str
            }
        elif dias_restantes < 0:
            dias_atraso = abs(dias_restantes)
            return {
                'status': f'ATRASADO ({dias_atraso} {"dia" if dias_atraso == 1 else "dias"})',
                'status_resumo': f'ATRASADO {dias_atraso}d',
                'recomendacao': f'⚠️ Atraso de {dias_atraso} {"dia" if dias_atraso == 1 else "dias"}. Cobrar transportadora.',
                'cor': CORES['ATRASADO'],
                'prioridade': 2,
                'previsao': previsao_str
            }
        elif dias_restantes <= 3:
            textos = {
                1: '🔴 ALERTA MÁXIMO – Previsão para amanhã!',
                2: '🟠 Atenção – 2 dias para o prazo final',
                3: '🟡 Pré-alerta – 3 dias para o prazo final'
            }
            return {
                'status': f'PREVISÃO VENCENDO ({dias_restantes} {"dia" if dias_restantes == 1 else "dias"})',
                'status_resumo': f'{dias_restantes}d',
                'recomendacao': textos.get(dias_restantes, f'Atenção: prazo em {dias_restantes} dias'),
                'cor': CORES[f'ALERTA_{dias_restantes}DIAS'],
                'prioridade': 3,
                'previsao': previsao_str
            }
        else:
            return {
                'status': f'NO PRAZO ({dias_restantes} {"dia" if dias_restantes == 1 else "dias restantes"})',
                'status_resumo': f'{dias_restantes}d',
                'recomendacao': '✅ Dentro do prazo – monitorar normalmente',
                'cor': CORES['PADRAO'],
                'prioridade': 4,
                'previsao': previsao_str
            }

    def consultar_pedido(self, chave: str) -> Dict:
        """Consulta a API SSW e processa a resposta"""
        self._inc('total_consultas')
        session = self._session()
        
        self._log(f"🔍 Consultando: {chave[:20]}...")
        
        try:
            # Usa POST com JSON (funciona perfeitamente)
            response = session.post(
                self.API_URL,
                json={'chave_nfe': chave},
                timeout=self.timeout
            )
            
            if response.status_code != 200:
                self._inc('total_erros')
                return self._resultado_erro(chave, f"HTTP {response.status_code}")
            
            # Tenta parsear como JSON
            try:
                dados = response.json()
            except:
                # Se não for JSON, tenta converter XML para dict
                try:
                    import xmltodict
                    dados = xmltodict.parse(response.text)
                except:
                    self._inc('total_erros')
                    return self._resultado_erro(chave, "Resposta não é JSON nem XML")
            
            # Verifica se a consulta foi bem sucedida
            success = dados.get('success', False)
            if not success:
                self._inc('total_sem_dados')
                return self._resultado_sem_dados(chave, dados.get('message', 'Documento não encontrado'))
            
            # Extrai os dados do documento
            documento = dados.get('documento', {})
            header = documento.get('header', {})
            
            # O campo pode ser 'tracking' (JSON) ou 'items' (XML convertido)
            tracking = documento.get('tracking', documento.get('items', []))
            
            if not tracking:
                self._inc('total_sem_dados')
                return self._resultado_sem_dados(chave, 'Sem eventos de rastreamento')
            
            # Classifica o status baseado nos eventos
            classificacao = self.classificar_status(tracking)
            
            # Último evento
            ultimo = tracking[-1]
            
            # Formata a data do último evento
            ultima_data = ultimo.get('data_hora', '')
            if ultima_data and 'T' in ultima_data:
                ultima_data = ultima_data.split('T')[0] + ' ' + ultima_data.split('T')[1][:5]
            
            resultado = {
                'nota_fiscal': header.get('nro_nf', ''),
                'numero_pedido': header.get('pedido', ''),
                'chave_nfe': chave,
                'destinatario': header.get('destinatario', ''),
                'remetente': header.get('remetente', ''),
                'status': classificacao['status'],
                'recomendacao': classificacao['recomendacao'],
                'prioridade': classificacao['prioridade'],
                'previsao': classificacao.get('previsao', ''),
                'data_entrega': classificacao.get('data_entrega', ''),
                'ultima_data': ultima_data,
                'ultima_situacao': ultimo.get('ocorrencia', ''),
                'ultimo_local': f"{ultimo.get('cidade', '')} - {ultimo.get('filial', '')}",
                'total_eventos': len(tracking),
                'data_consulta': datetime.now().strftime('%d/%m/%Y %H:%M')
            }
            
            self._inc('total_sucessos')
            self._log(f"   ✅ {resultado['status']}")
            
            # Log especial para devolução
            if 'DEVOLVIDO' in resultado['status']:
                self._log(f"   🔄 ATENÇÃO: Devolução detectada!", "aviso")
            
            return resultado
            
        except requests.exceptions.Timeout:
            self._inc('total_erros')
            return self._resultado_erro(chave, "Timeout na requisição")
        except Exception as e:
            self._inc('total_erros')
            self._log(f"   ❌ Erro: {e}", "erro")
            return self._resultado_erro(chave, str(e)[:100])

    # ------------------------------------------------------------------
    # PROCESSAMENTO PARALELO
    # ------------------------------------------------------------------

    def processar_lote(
        self,
        df: pd.DataFrame,
        max_consultas: int = None,
        callback_progresso=None
    ) -> pd.DataFrame:
        """Processa múltiplos pedidos em paralelo"""
        
        if max_consultas:
            df = df.head(max_consultas)
        
        total = len(df)
        self._log(f"\n{'='*70}")
        self._log(f"🚀 PROCESSANDO {total} PEDIDOS | Workers: {self.workers}")
        self._log(f"{'='*70}\n")
        
        # Prepara tarefas
        tarefas: List[tuple] = []
        for idx, row in df.iterrows():
            chave = row['chave_nfe']
            extra = {c: row[c] for c in df.columns if c != 'chave_nfe'}
            tarefas.append((idx, chave, extra))
        
        resultados_ordenados = [None] * total
        progresso = {'atual': 0}
        inicio = time.time()
        
        with ThreadPoolExecutor(max_workers=self.workers) as executor:
            future_to_pos = {
                executor.submit(self.consultar_pedido, chave): (pos, chave, extra)
                for pos, (idx, chave, extra) in enumerate(tarefas)
            }
            
            for future in as_completed(future_to_pos):
                pos, chave, extra = future_to_pos[future]
                try:
                    resultado = future.result()
                except Exception as e:
                    resultado = self._resultado_erro(chave, str(e))
                
                # Preserva colunas extras
                for col, val in extra.items():
                    if col not in resultado:
                        resultado[col] = val
                
                resultados_ordenados[pos] = resultado
                
                with self._lock:
                    progresso['atual'] += 1
                    atual = progresso['atual']
                
                if callback_progresso:
                    try:
                        callback_progresso(atual, total, resultado)
                    except Exception:
                        pass
                
                # Log de progresso a cada 10% ou 20 itens
                if atual % max(1, min(20, total // 5)) == 0 or atual == total:
                    elapsed = time.time() - inicio
                    taxa = atual / elapsed if elapsed > 0 else 0
                    restante = (total - atual) / taxa if taxa > 0 else 0
                    self._log(f"📊 Progresso: [{atual}/{total}] - {taxa:.1f} req/s - ~{int(restante)}s restantes")
        
        elapsed_total = time.time() - inicio
        self._log(f"\n✅ PROCESSAMENTO CONCLUÍDO em {elapsed_total:.1f} segundos")
        self._log(f"   Média: {total/elapsed_total:.1f} consultas/segundo\n")
        
        return pd.DataFrame(resultados_ordenados)

    # ------------------------------------------------------------------
    # HELPERS DE RETORNO
    # ------------------------------------------------------------------

    def _resultado_erro(self, chave: str, motivo: str) -> Dict:
        return {
            'nota_fiscal': '', 
            'numero_pedido': '', 
            'chave_nfe': chave,
            'destinatario': '', 
            'remetente': '',
            'status': 'ERRO NA CONSULTA',
            'recomendacao': f'❌ Erro: {motivo}. Verificar manualmente.',
            'prioridade': 1, 
            'previsao': '', 
            'data_entrega': '',
            'ultima_data': '', 
            'ultima_situacao': 'Erro ao consultar API',
            'ultimo_local': '', 
            'total_eventos': 0,
            'data_consulta': datetime.now().strftime('%d/%m/%Y %H:%M')
        }

    def _resultado_sem_dados(self, chave: str, mensagem: str = "") -> Dict:
        return {
            'nota_fiscal': '', 
            'numero_pedido': '', 
            'chave_nfe': chave,
            'destinatario': '', 
            'remetente': '',
            'status': 'AGUARDANDO RASTREIO',
            'recomendacao': '📭 Sem rastreio na SSW. Verificar manualmente.',
            'prioridade': 4, 
            'previsao': '', 
            'data_entrega': '',
            'ultima_data': '', 
            'ultima_situacao': mensagem or 'Nenhum dado encontrado',
            'ultimo_local': '', 
            'total_eventos': 0,
            'data_consulta': datetime.now().strftime('%d/%m/%Y %H:%M')
        }

    # ------------------------------------------------------------------
    # GERAÇÃO DE RELATÓRIOS
    # ------------------------------------------------------------------

    def gerar_relatorios(self, df: pd.DataFrame, nome_base: str):
        """Gera relatórios Excel e CSV formatados - VERSÃO OTIMIZADA PARA MEMÓRIA"""
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        
        # ============================================
        # 1. ESTATÍSTICAS (antes de modificar o df)
        # ============================================
        entregues = df[df['status'].str.contains('ENTREGUE', na=False)].shape[0]
        devolvidos = df[df['status'].str.contains('DEVOLVIDO', na=False)].shape[0]
        atrasados = df[df['status'].str.contains('ATRASADO', na=False)].shape[0]
        alerta = df[df['status'].str.contains('PREVISÃO VENCENDO', na=False)].shape[0]
        aguardando = df[df['status'].str.contains('AGUARDANDO', na=False)].shape[0]
        erros = df[df['status'].str.contains('ERRO', na=False)].shape[0]
        
        # Em trânsito = todos menos os classificados
        em_transito = len(df) - entregues - devolvidos - atrasados - alerta - aguardando - erros
        
        self._log(f"\n{'='*70}")
        self._log(f"📊 RESUMO FINAL DO RASTREAMENTO:")
        self._log(f"{'='*70}")
        self._log(f"   ✅ ENTREGUES:        {entregues}")
        self._log(f"   🔄 DEVOLVIDOS:       {devolvidos}")
        self._log(f"   🚚 EM TRÂNSITO:      {em_transito}")
        self._log(f"   ⚠️ ATRASADOS:        {atrasados}")
        self._log(f"   🟡 PREVISÃO VENCENDO: {alerta}")
        self._log(f"   📭 AGUARDANDO:       {aguardando}")
        self._log(f"   ❌ ERROS:            {erros}")
        self._log(f"{'='*70}\n")
        
        # ============================================
        # 2. PREPARA O DATAFRAME (remove colunas desnecessárias)
        # ============================================
        ordem_colunas = [
            'nota_fiscal', 'numero_pedido', 'destinatario',
            'status', 'recomendacao', 'previsao', 'data_entrega',
            'ultima_data', 'ultima_situacao', 'ultimo_local',
            'total_eventos', 'data_consulta'
        ]
        
        # Adiciona colunas extras que existirem
        for col in df.columns:
            if col not in ordem_colunas and col not in ['chave_nfe', 'remetente', 'prioridade']:
                ordem_colunas.append(col)
        
        colunas_existentes = [c for c in ordem_colunas if c in df.columns]
        df_rel = df[colunas_existentes].copy()
        
        # Libera o DataFrame original da memória (opcional, ajuda o GC)
        del df
        
        # ============================================
        # 3. GERA EXCEL USANDO OPENPYXL (LINHA POR LINHA)
        # ============================================
        from openpyxl import Workbook
        from openpyxl.styles import PatternFill, Font, Alignment
        from openpyxl.utils import get_column_letter
        
        arquivo = f"{nome_base}_rastreamento_{timestamp}.xlsx"
        
        try:
            # Cria workbook
            wb = Workbook()
            ws = wb.active
            ws.title = 'Rastreamento'
            
            # Escreve cabeçalhos
            header_fill = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
            header_font = Font(bold=True, color='FFFFFF')
            
            for col_idx, header in enumerate(colunas_existentes, 1):
                cell = ws.cell(row=1, column=col_idx, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center')
            
            # Mapeia cores por status (pré-computado para performance)
            def get_cor_status(status):
                if 'ENTREGUE' in status:
                    return CORES['ENTREGUE']
                elif 'DEVOLVIDO' in status:
                    return CORES['DEVOLVIDO']
                elif 'ATRASADO' in status:
                    return CORES['ATRASADO']
                elif 'PREVISÃO VENCENDO' in status:
                    if '(1' in status or '1 dia' in status:
                        return CORES['ALERTA_1DIA']
                    elif '(2' in status or '2 dias' in status:
                        return CORES['ALERTA_2DIAS']
                    else:
                        return CORES['ALERTA_3DIAS']
                elif 'ERRO' in status:
                    return CORES['ERRO']
                elif 'AGUARDANDO' in status:
                    return CORES['SEM_DADOS']
                else:
                    return CORES['PADRAO']
            
            # ⭐ CRÍTICO: Escreve linha por linha (NÃO carrega tudo na memória)
            linha_atual = 2
            total_linhas = len(df_rel)
            
            for idx, row in enumerate(df_rel.itertuples(index=False), start=2):
                # Escreve os valores da linha
                for col_idx, value in enumerate(row, 1):
                    ws.cell(row=linha_atual, column=col_idx, value=value)
                
                # Aplica cor baseada no status (pega a primeira coluna ou a coluna 'status')
                status = str(row[colunas_existentes.index('status')] if 'status' in colunas_existentes else '')
                cor_fundo = get_cor_status(status)
                
                if cor_fundo != CORES['PADRAO']:
                    fill = PatternFill(start_color=cor_fundo, end_color=cor_fundo, fill_type='solid')
                    for col_idx in range(1, len(colunas_existentes) + 1):
                        ws.cell(row=linha_atual, column=col_idx).fill = fill
                
                linha_atual += 1
                
                # A cada 500 linhas, libera um pouco de memória
                if linha_atual % 500 == 0:
                    self._log(f"   Gerando Excel: {linha_atual - 1}/{total_linhas} linhas...", "debug")
            
            # Ajusta largura das colunas
            for col in ws.columns:
                max_len = 0
                col_letter = get_column_letter(col[0].column)
                for cell in col:
                    try:
                        if cell.value:
                            max_len = max(max_len, len(str(cell.value)))
                    except:
                        pass
                worksheet.column_dimensions[col_letter].width = min(max_len + 2, 50)
            
            # Salva o arquivo
            wb.save(arquivo)
            wb.close()
            
            self._log(f"✅ Relatório Excel gerado: {arquivo}")
            
        except Exception as e:
            self._log(f"❌ Erro ao gerar Excel: {e}", "erro")
            # Fallback: tenta gerar CSV se Excel falhar
            arquivo = self._gerar_csv_fallback(df_rel, nome_base, timestamp)
        
        # ============================================
        # 4. GERA CSV PARA INTELIPOST (se houver entregues)
        # ============================================
        try:
            if entregues > 0:
                arquivo_csv = f"{nome_base}_intelipost_{timestamp}.csv"
                
                # Filtra apenas entregues (cria cópia pequena)
                entregues_mask = df_rel['status'].str.contains('ENTREGUE', na=False)
                
                if entregues_mask.any():
                    # Encontra coluna de pedido
                    col_pedido = None
                    for col in ['numero_pedido', 'nota_fiscal', 'chave_nfe']:
                        if col in df_rel.columns and df_rel[col].notna().any():
                            col_pedido = col
                            break
                    
                    if col_pedido:
                        # Cria DataFrame pequeno só com entregues
                        entregues_df = df_rel.loc[entregues_mask, [col_pedido, 'data_entrega']].copy()
                        entregues_df['status'] = 'ENTREGUE'
                        entregues_df.to_csv(arquivo_csv, index=False, encoding='utf-8-sig')
                        self._log(f"✅ CSV Intelipost gerado: {arquivo_csv}")
                        
                        # Libera memória
                        del entregues_df
        except Exception as e:
            self._log(f"⚠️ Erro ao gerar CSV Intelipost: {e}", "aviso")
        
        # ============================================
        # 5. GERA RELATÓRIO DE DEVOLUÇÕES
        # ============================================
        try:
            if devolvidos > 0:
                arquivo_dev = f"{nome_base}_devolucoes_{timestamp}.csv"
                devolvidos_mask = df_rel['status'].str.contains('DEVOLVIDO', na=False)
                
                if devolvidos_mask.any():
                    devolvidos_df = df_rel.loc[devolvidos_mask].copy()
                    devolvidos_df.to_csv(arquivo_dev, index=False, encoding='utf-8-sig')
                    self._log(f"⚠️ Relatório de devoluções gerado: {arquivo_dev}")
                    
                    # Libera memória
                    del devolvidos_df
        except Exception as e:
            self._log(f"⚠️ Erro ao gerar relatório de devoluções: {e}", "aviso")
        
        # Libera o DataFrame final da memória
        del df_rel
        
        return arquivo


    def _gerar_csv_fallback(self, df_rel: pd.DataFrame, nome_base: str, timestamp: str) -> str:
        """Fallback: gera CSV quando o Excel falha por falta de memória"""
        arquivo = f"{nome_base}_rastreamento_{timestamp}.csv"
        self._log(f"⚠️ Gerando CSV como fallback (menos memória): {arquivo}", "aviso")
        
        # Salva CSV em chunks para economizar memória
        chunk_size = 500
        total_rows = len(df_rel)
        
        # Primeiro chunk com cabeçalho
        df_rel.head(chunk_size).to_csv(arquivo, index=False, encoding='utf-8-sig')
        
        # Demais chunks em modo append
        for start in range(chunk_size, total_rows, chunk_size):
            df_rel.iloc[start:start + chunk_size].to_csv(
                arquivo, mode='a', header=False, index=False, encoding='utf-8-sig'
            )
            self._log(f"   CSV: {min(start + chunk_size, total_rows)}/{total_rows} linhas...", "debug")
        
        self._log(f"✅ Relatório CSV gerado (fallback): {arquivo}")
        return arquivo


# ============================================
# TESTE RÁPIDO
# ============================================

def testar_chave_unica():
    """Função para testar uma única chave"""
    print("\n" + "="*70)
    print(" TESTE DE CONSULTA SSW - VERSÃO FINAL")
    print("="*70)
    
    chave = input("\nDigite a chave NF-e (44 dígitos): ").strip()
    chave = re.sub(r'[^0-9]', '', chave)
    
    if len(chave) != 44:
        print(f"❌ Chave inválida! Tem {len(chave)} dígitos, deveria ter 44")
        return
    
    processador = ProcessadorSSW(workers=1)
    print("\n🔍 Consultando API...\n")
    
    resultado = processador.consultar_pedido(chave)
    
    print("\n" + "="*70)
    print(" RESULTADO DA CONSULTA:")
    print("="*70)
    for key, value in resultado.items():
        if value:
            print(f"   {key}: {value}")
    print("="*70)
    
    # Salva resultado em JSON
    with open('consulta_resultado.json', 'w', encoding='utf-8') as f:
        json.dump(resultado, f, ensure_ascii=False, indent=2)
    print("\n📄 Resultado salvo em: consulta_resultado.json")


if __name__ == "__main__":
    testar_chave_unica()